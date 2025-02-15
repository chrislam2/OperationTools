import os
import re
import shutil
import openpyxl
import datetime
import getpass

###############################
# Purposes
# Export a report showing the reference paths in the code files
#
# Note: 
# 1. Please change the global variables to assign the checking and the report location
# 2. Do not include .xlsx and .xlsm
###############################

###############################
# Global Variables
###############################
login_name = os.getlogin()
cUser = os.path.expanduser('~')
current_directory = os.path.dirname(os.path.realpath(__file__))
previous_directory = os.path.abspath(os.path.join(current_directory, os.pardir))

input_folder = 'C:\\Users\\'
output_txt_file = os.path.join(current_directory, "exported_paths.txt")
output_xlsx_file = os.path.join(current_directory, "exported_paths.xlsx")

file_extensions = ['.R', '.py', '.vbs', '.bat', '.ps1']
specific_search = ['.csv', '.xlsx', '.xlsm', '.txt', '.gz', '.html']

###############################
# Functions
###############################
def find_reference_paths(file_path, specific_search=['ndams-sspp01']):
    reference_paths_with_lines = []
    line_number_occurence = []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            for line_number, line in enumerate(file, start=1):
                paths = []
                paths.extend(re.findall(r'[a-zA-Z]:[\\/]{1,2}(?:[^\\/:*?"<>|\r\n{}]+[\\/]{1,2})*[^\\/:*?"<>|\r\n{}]*(?=\s|$|#|;|\'|")', line))
                paths.extend(re.findall(r'(?:[a-zA-Z]:[\\/]{1,2}|\\\\[\w.-]+\\[\w\s.-]+)(?:[^\\/:*?"<>|\r\n{}]+[\\/]{1,2})*[^\\/:*?"<>|\r\n{}]*(?=\s|$|#|;|\'|")', line))
                paths.extend(re.findall(r'(?:[a-zA-Z]:[\\/]{1,2}|\\\\[\w.-]+\\[\w\s.-]+)\S*(?=\s|$|#|;|\'|")', line))
                if any(keyword in line for keyword in specific_search):
                    paths.append(line)
                for path in paths:
                    if line_number not in line_number_occurence:
                        line_number_occurence.append(line_number)
                        reference_paths_with_lines.append((line_number, line.rstrip()))
    except Exception as e:
        print(e)
    return reference_paths_with_lines

def export_reference_paths_to_txt(reference_paths_dict, output_file):
    with open(output_file, 'w', encoding='utf-8') as file:
        for file_name, (file_path, paths) in reference_paths_dict.items():
            file.write(file_name + "\n")
            file.write("file path: " + file_path + "\n")
            for line_number, line in paths:
                file.write("line " + str(line_number) + ": " + line + "\n")
            file.write("\n")

def export_reference_paths_to_xlsx(reference_paths_dict, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['File Name', 'File Path', 'Last Modified Date', 'Last Modified By', 'Line Number', 'Line Content'])
    prev_file_name = None
    prev_file_path = None
    for file_name, (file_path, paths) in reference_paths_dict.items():
        for line_number, line in paths:
            if file_name == prev_file_name and file_path == prev_file_path:
                ws.append(['', '', '', '', line_number, line])
            else:
                ws.append([file_name, file_path, datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'), getpass.getuser(), line_number, line])
                prev_file_name = file_name
                prev_file_path = file_path
    wb.save(output_file)

def copy_files_from_excel(input_xlsx_file):
    wb = openpyxl.load_workbook(input_xlsx_file)
    ws = wb.active

    copy_occurence = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        file_name, file_path, line_number, line_content, new_file_path, new_content = row

        if new_file_path:
            if os.path.exists(file_path):
                if os.path.abspath(file_path) == os.path.abspath(new_file_path):
                    base_name, extension = os.path.splitext(file_name)
                    new_file_name = base_name + "_copy" + extension
                    new_file_path = os.path.join(os.path.dirname(file_path), new_file_name)
                    print("Destination path is the same as the source. Copying to " + new_file_path + " instead.")
                if new_file_path not in copy_occurence:
                    shutil.copy(file_path, new_file_path)
                    print("File " + file_name + " copied to " + new_file_path + " successfully.")
                    copy_occurence.append(new_file_path)
                else:
                    print(new_file_path + " has already been copied.")
            else:
                print("File " + file_name + " does not exist at " + file_path + ". Skipping copy.")

    wb.close()

def replace_line_content(input_xlsx_file):
    wb = openpyxl.load_workbook(input_xlsx_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        file_name, file_path, line_number, line_content, new_file_path, new_content = row
        if new_content:
            if os.path.abspath(file_path) == os.path.abspath(new_file_path):
                base_name, extension = os.path.splitext(file_name)
                new_file_name = base_name + "_copy" + extension
                new_file_path = os.path.join(os.path.dirname(file_path), new_file_name)
            with open(new_file_path, 'r', encoding='utf-8', errors='ignore') as file:
                lines = file.readlines()

            if len(lines) >= line_number:
                existing_line_content = lines[line_number - 1].strip()
                
                if existing_line_content == line_content:
                    lines[line_number - 1] = new_content + '\n'

                    with open(new_file_path, 'w', encoding='utf-8') as file:
                        file.writelines(lines)
                    print("Replacement done for " + file_name + " in " + new_file_path + ", line " + str(line_number))
                else:
                    print("Line content in " + file_name + " in " + new_file_path + ", line " + str(line_number) + " doesn't match. Skipped replacement.")

    wb.close()

###############################
# Main Process
###############################
file_paths_dict = {}
for root, dirs, files in os.walk(input_folder):
    for file in files:
        if any(file.endswith(ext) for ext in file_extensions):
            file_path = os.path.join(root, file)
            print("Scanning files: " + file_path)
            file_paths = find_reference_paths(file_path, specific_search)
            if file_paths:
                file_paths_dict[file] = (file_path, file_paths)

if file_paths_dict:
    export_reference_paths_to_txt(file_paths_dict, output_txt_file)
    print("File paths exported to the txt file: " + output_txt_file)
    export_reference_paths_to_xlsx(file_paths_dict, output_xlsx_file)
    print("File paths exported to the xlsm file: " + output_xlsx_file)
else:
    print("No file paths found in the documents.")